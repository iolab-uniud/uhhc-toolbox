
from pydantic import ValidationError
from .. models import Instance, Solution
from .. models.instance_models import CostComponents, Synchronization, Instance, Caregiver, Patient, Service, RequiredService, TerminalPoint, MetaData

import re
import sys
import ast
import json
import click
import openpyxl
import warnings
import pandas as pd
import configparser

@click.group()
def cli():
    """This is the validator group"""
    pass

@cli.group()
def validator():
    """Commands related to instance and solution validator"""
    pass

@validator.command()
@click.argument('filename', type=click.File())
def instance(filename):
    try:
        c = Instance.model_validate_json(filename.read())    
        click.secho(f"Validation OK, instance signature: {c.signature}", fg='green')
    except ValidationError as e:
        click.secho(f"{e}", fg='red')


@validator.command()
@click.argument('instance-filename', type=click.File())
@click.argument('solution-filename', type=click.File())
def solution(instance_filename, solution_filename):
    try:
        i = Instance.model_validate_json(instance_filename.read())
        s = Solution.model_validate_json(solution_filename.read())    
        click.secho(f"Solution uploaded. About to check validity", fg='green')
        try:
            s.check_validity(i)
        except Exception as e:
            click.secho(f"{e}", fg="red")
            click.secho(f"Not continuing validation. Exiting validator...", fg="red")
            sys.exit(1)
            # raise ValidationError(e)
        click.secho(f"Validity checked. About to compute costs.", fg='green')
        s.compute_costs(i)
        # click.secho("Costs are: ", fg="green")
        # max_key_length = max(len(key) for key in data)
        # max_value_number_length = max(len(value.split()[0]) for value in data.values())

        # for key, value in data.items():
        #     num, rest = value.split(maxsplit=1)
        #     formatted_line = f"{key.ljust(max_key_length + 5, '.')} {num.rjust(max_value_number_length)} ..... {rest}"
        #     print(formatted_line)
        #print(json.dumps(costs,sort_keys=True, indent=2))
    except ValidationError as e:
        click.secho(f"{e}", fg='red')

@cli.group()
def utils():
    """Utilities"""
    pass

@utils.command()
@click.argument('filename', type=click.File())
@click.option('--format', type=click.Choice(['json', 'latex']), default='json', help="Output format")
@click.option('--pretty', is_flag=True, show_default=True, default=False, help="Pretty print the output")
def instance_features(filename, format, pretty):
    c = Instance.model_validate_json(filename.read())
    if format == 'json':
        if not pretty:
            print(json.dumps(c.features))
        else:
            class RoundingFloat(float):
                __repr__ = staticmethod(lambda x: f'{x:.3f}')
            json.encoder.c_make_encoder = None
            json.encoder.float = RoundingFloat
            print(json.dumps(c.features, indent=3))
    elif format == 'latex':        
        reformed_dict = {} 
        for outerKey, innerDict in c.features.items(): 
            if type(innerDict) == dict:
                for innerKey, values in innerDict.items(): 
                    reformed_dict[(outerKey, innerKey)] = [values]
            else:
                reformed_dict[(outerKey, )] = [innerDict]
        df = pd.DataFrame(reformed_dict)
        print(df.to_latex(index=False))

    
@utils.command()
@click.argument('instance-filename', type=click.File())
@click.argument('solution-filename', type=click.File())
@click.option('--output', '-o', type=click.Path(dir_okay=False, writable=True), help="Output file")
def plot_solution(instance_filename, solution_filename, output):
    from .plot import plot
    i = Instance.model_validate_json(instance_filename.read())
    s = Solution.model_validate_json(solution_filename.read())    
    s.check_validity(i)
    plt = plot(i, s)
    if not output:
        plt.show()
    else:
        plt.tight_layout()
        plt.savefig(output, bbox_inches='tight')

@utils.command()
@click.argument('instance-filename', type=click.File())
@click.argument('solution-filename', type=click.File())
@click.option('--output', '-o', type=click.Path(dir_okay=False, writable=True), help="Output file")
def view_solution(instance_filename, solution_filename, output):
    from .interactive_visualizer import visualize
    i = Instance.model_validate_json(instance_filename.read())
    s = Solution.model_validate_json(solution_filename.read())    
    s.check_validity(i)
    fig = visualize(i, s)
    if not output:
        fig.show()
    else:
        assert output.endswith('.html'), "Output file must be an HTML file"
        fig.write_html(output)

@utils.command()
@click.argument('what', type=click.Choice(['instance', 'solution']), default='instance')
def show_schema(what):
    if what == 'instance':
        print(json.dumps(Instance.model_json_schema(), indent=2))
    elif what == 'solution':
        print(json.dumps(Solution.model_json_schema(), indent=2))

@utils.command()
@click.argument('workbook-path', type=str)
@click.argument('sheet-to-consider', type=int)
def convert_instance_bazirha(workbook_path, sheet_to_consider):
    workbook_name = workbook_path.split("/")[-1].split(".")[0]
    time_window_met = "at_service_end" # This is from the original paper
    name_instance = f"{workbook_name}{sheet_to_consider+1}" 
    name_depot_departure = "d1"
    horizon = 1000
    origin = "bazirha-caie"
    if workbook_name in ["A","B","C","D","E","F"]:
        origin = "bazirha"
    def find_subtable_indices(sheet, header_pattern):
        pattern = re.compile(header_pattern, re.IGNORECASE)  # Case-insensitive regex pattern
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value and pattern.match(str(cell.value).strip()):
                    return cell.row, cell.column
        return None, None

    def extract_data(sheet, start_row, start_col, end_col=None, is_serv_time=False, index=True):
        rows = []
        for row in sheet.iter_rows(min_row=start_row):
            if row[0].value is None and len(rows) > 0:  # Assuming empty row indicates end of subtable
                if is_serv_time: # This is a bit silly: The table related to the sevice time is the only one with a different format as one as also the center
                    if row[start_col].value is None:
                        break
                else:
                    break
            if end_col is None:
                row_data = [cell.value if type(cell.value) != str else cell.value.strip() for cell in row[start_col-1:]]
            else:
                row_data = [cell.value if type(cell.value) != str else cell.value.strip() for cell in row[start_col-1:end_col-1]]
            rows.append(row_data)
        df = pd.DataFrame(rows[1:])
        df.columns = map(lambda s: s.strip() if type(s) == str else s, rows[0])
        df = df.dropna(axis='columns', how='all').dropna(axis='rows', how='all')
        return df.set_index(df.columns[0]) if index else df
    workbook = openpyxl.load_workbook(workbook_path, data_only=True)
    # Data extraction for each sheet
    all_data = {}
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        # Find indices of each header
        pat_idx, pat_col = find_subtable_indices(sheet, r"Patients")
        serv_idx, serv_col = find_subtable_indices(sheet, r"Pat\s*/\s*Ser")
        serv_time_idx, serv_time_col = find_subtable_indices(sheet, r"P/SerTime")
        skill_idx, skill_col = find_subtable_indices(sheet, r"CgSkills")
        tw_caregivers_idx, tw_caregivers_col = find_subtable_indices(sheet, r"TWCaregivers")
        tw_patients_idx, tw_patients_col = find_subtable_indices(sheet, r"Tw\s+Patients")

        assert (pat_idx, pat_col) != (None, None)
        assert (serv_idx, serv_col) != (None, None)
        assert (serv_time_idx, serv_time_col) != (None, None)
        assert (skill_idx, skill_col) != (None, None)
        assert (tw_caregivers_idx, tw_caregivers_col) != (None, None)
        assert (tw_patients_idx, tw_patients_col) != (None, None)

        # Extract each data section    
        patients_df = extract_data(sheet, pat_idx, pat_col)
        services_df = extract_data(sheet, serv_idx, serv_col, serv_time_col)
        service_times_df = extract_data(sheet, serv_time_idx, serv_time_col, skill_col, True)
        skills_df = extract_data(sheet, skill_idx, skill_col, tw_caregivers_col)
        tw_caregivers_df = extract_data(sheet, tw_caregivers_idx, tw_caregivers_col, tw_patients_col)
        tw_patients_df = extract_data(sheet, tw_patients_idx, tw_patients_col)
        all_data[sheet_name] = {
            "Patients & Travel Times": patients_df,
            "Service Requirements": services_df,
            "Service Times": service_times_df,
            "Caregiver Skills": skills_df,
            "Time Windows Caregivers": tw_caregivers_df,
            "Time Windows Patients": tw_patients_df
        }
    service_times_df = extract_data(sheet, serv_time_idx, serv_time_col, skill_col)
    patients_df = all_data[workbook.sheetnames[sheet_to_consider]]["Patients & Travel Times"]
    services_df = all_data[workbook.sheetnames[sheet_to_consider]]["Service Requirements"]
    service_times_df = all_data[workbook.sheetnames[sheet_to_consider]]["Service Times"]
    skills_df = all_data[workbook.sheetnames[sheet_to_consider]]["Caregiver Skills"]
    tw_caregivers_df = all_data[workbook.sheetnames[sheet_to_consider]]["Time Windows Caregivers"]
    tw_patients_df = all_data[workbook.sheetnames[sheet_to_consider]]["Time Windows Patients"]

    services = {}

    for service in services_df.columns:
        if "Service" not in service:
            continue
        s = f"s{service.split()[-1]}"
        services[s] = [Service(id=s, type=s), 0, 0] 

    def dispatch_patients(patients_df, services_df, service_times_df, tw_patients_df):
        patients = []
        for i in range(1, patients_df.shape[0] - 1):
            patient_row = patients_df.iloc[i]
            patient_id = patients_df.index[i]
            assert patient_id == i, f"Misplaced patient, expecting {i} found {patient_id}"
            
            assert patient_id  == int(str(services_df.iloc[i - 1] .name).strip().split()[-1]), f"Misplaces patient in services_df, expecting {patient_id}, found {str(services_df.iloc[i - 1] .name).strip().split()[-1]}"
            assert patient_id  == int(str(service_times_df.iloc[i] .name).strip().split()[-1]), f"(This check is actually not necessary) Misplaces patient in service_times_df, expecting {patient_id}, found {str(services_df.iloc[i] .name).strip().split()[-1]}"
            patient_services = list(filter(lambda s: not s.startswith('Lamda'), services_df.columns[services_df.iloc[i - 1] == 1]))  
            # patient_services = [f"s{s.split()[-1]}" for s in patient_services_raw]
            required_services = []    
            for raw_s, t in service_times_df.loc[f"Patient {patient_id}", patient_services].items():  
                s = f"s{raw_s.split()[-1]}"
                required_services.append(RequiredService(service=s, duration=t))
                if s not in services:
                    services[s] = [Service(id=s, type=s), 1, t] 
                else:
                    services[s][1] += 1
                    services[s][2] += t  
            
            location = (float(patient_row["X"]), float(patient_row["Y"]))

            tw_patient_row = tw_patients_df.iloc[i - 1]
            assert patient_id  == int(str(tw_patient_row.name).strip().split()[-1]), f"Misplaces patient in tw_patients_df, expecting {patient_id}, found {str(tw_patient_row.name).strip().split()[-1]}"
            # Here it should be a bit more principled.
            # You do not know how many TW are there, BUT you can have an hint on that based on the shape of the related df
            time_windows_patient = []
            for c_tw in range(int(tw_patients_df.shape[1] / 2)):
                tw_start = tw_patient_row[f"ETW {c_tw + 1}"] 
                tw_end = tw_patient_row[f"LTW {c_tw + 1}"] 
                time_windows_patient.append((tw_start,tw_end))

            type_of_synchronization = "simultaneous"
            simultaneous = 0
            if "Lamda_i" in list(services_df.columns):
                simultaneous = services_df.iloc[i - 1]["Lamda_i"] 
            else:
                print("Lambda i not there")
            if not simultaneous:
                type_of_synchronization = "independent"

            synch = Synchronization(type=type_of_synchronization)
            if len(required_services) < 2:
                synch = None
            
            p = Patient(id=f"p{patient_id}", 
                        required_services=required_services, 
                        distance_matrix_index=patient_id, 
                        time_windows=time_windows_patient,
                        synchronization=synch, 
                        location=location)
            patients.append(p)
            #required_services = { s: t for s, t in service_times_df[patient_services] }
            #print(patient_id, required_services)
        return patients

    def dispatch_departing_points(patients_df):
        # This is a list just because the Instance class/method the require a list
        departing_points = [] # list that containts DepartingPoints, each departing point is, in this case the center (index 0 in patients)
        center_id = 0
        center_row = patients_df.iloc[center_id]
        location = (float(center_row["X"]), float(center_row["Y"]))
        dp = TerminalPoint(
            id=name_depot_departure,
            distance_matrix_index=center_id,
            location=location
        )
        departing_points.append(dp)
        return departing_points

    # def dispatch_arrival_points(patients_df):
    #     # This is a list just because the Instance class/method the require a list
    #     arrival_points = []
    #     arrival_point_id = patients_df.shape[0]-1
    #     arrival_point_row = patients_df.iloc[arrival_point_id]
    #     location = (float(arrival_point_row["X"]), float(arrival_point_row["Y"]))
    #     ap = TerminalPoint(
    #         id=name_depot_arrival,
    #         distance_matrix_index=arrival_point_id,
    #         location=location
    #     )
    #     arrival_points.append(ap)
    #     return arrival_points

    def dispatch_caregivers(tw_caregivers_df,skills_df):
        caregivers = [] 
        for i in range(tw_caregivers_df.shape[0]):
            caregiver_id = f"c{i + 1}" # This is because range goew from 0 to n-1 while caregiver_ids start from 1
            
            # check the caregivers are in order in the skills df
            assert i + 1 ==  int(str(skills_df.iloc[i].name).strip().split()[-1]), f"Caregiver misplaced in tw_caregiver_df, expected {i+1} found {str(skills_df.iloc[i].name).strip().split()[-1]}"
            caregiver_services_raw = list(skills_df.columns[skills_df.iloc[i] == 1]) 
            # give the services the "correct" name
            caregiver_services = [f"s{s.split()[-1]}" for s in caregiver_services_raw]
            for s in caregiver_services:
                if s not in services.keys():
                    services[s] = [Service(id=s, type=s), 0, 0] 
            
            assert i + 1 ==  int(str(tw_caregivers_df.iloc[i].name).strip().split()[-1]), f"Caregiver misplaced in tw_caregivers_df, expected {i+1} found {str(tw_caregivers_df.iloc[i].name).strip().split()[-1]}"
            tw_s = int(tw_caregivers_df.iloc[i]["ETW"])
            tw_e = int(tw_caregivers_df.iloc[i]["LTW"])
            
            c = Caregiver(id=caregiver_id,
                        abilities=caregiver_services,
                        departing_point=name_depot_departure,
                        arrival_point=name_depot_departure, 
                        working_shift=(tw_s,tw_e))
            caregivers.append(c)
        return caregivers 

    def dispatch_services(services):
        generated_services = []
        for s_key in services.keys():
            s = services[s_key][0]
            generated_services.append(s)
        return generated_services

    def dispatch_distances(patients_df):
        distances = []
        separator_col = patients_df.columns.get_loc("Travel time")
        for i in  range(patients_df.shape[0]): 
            d = list(patients_df.iloc[i, separator_col + 1:-1]) # get from the column after the column separator to the end, considering row i
            distances.append(d)
        return distances[:-1]
    
    generated_patients = dispatch_patients(patients_df, services_df, service_times_df, tw_patients_df)
    generated_departing_points = dispatch_departing_points(patients_df)
    generated_caregivers = dispatch_caregivers(tw_caregivers_df, skills_df)
    generated_services = dispatch_services(services)
    generated_distances = dispatch_distances(patients_df)
    cc = CostComponents() 
    if workbook_name in ["A","B","C","D","E","F"]:
        cc = CostComponents(travel_time=1, 
                        total_tardiness="HARD", 
                        total_extra_time="HARD"
        )
    else:
        cc = CostComponents(
            total_tardiness="HARD", 
            total_extra_time="HARD",
            total_waiting_time=1,
            workload_balance = 1
        )

        # "cost_components": {
        #   "total_tardiness": "HARD",
        #   "total_extra_time": "HARD",
        #   "total_waiting_time": 1,
        #   "workload_balance": 1
        # },
    

    metadata = MetaData(time_window_met=time_window_met, 
                        cost_components=cc,
                        name=name_instance,
                        horizon=horizon,
                        origin=origin)
    
    instance = Instance(
            metadata = metadata,
            distances=generated_distances, 
            terminal_points=generated_departing_points,
            caregivers=generated_caregivers,
            patients=generated_patients,
            services=generated_services,
        )

    with open(f"{instance.metadata.name}.json", "w") as f:
            f.write(instance.model_dump_json(indent=2, exclude_none=True))

@utils.command()
@click.argument('instance-filename', type=str)
@click.argument('output', type=str)
@click.option('--tau', '-t', type=int, help="Time multiplier", default=5)
def convert_instance_raidl(instance_filename, output, tau):

    car_matrix_file = f"{instance_filename}.CAR.dm"
    public_matrix_file = f"{instance_filename}.PUBLIC.dm"
    entities_file = f"{instance_filename}.conf"

    terminal_points = set()
    services = set()

    def parse_entities_file(filename):
        config = configparser.ConfigParser()
        config.read(filename)
        
        normalization = {}
        nurses = []
        customers = []
        jobs = []
        
        for section in config.sections():
            data = dict(config.items(section))
            
            # Convert numerical values where applicable
            for key, value in data.items():
                try:
                    data[key] = ast.literal_eval(value)
                except (ValueError, SyntaxError):
                    pass
            
            if section == "normalizations":
                normalization = data
            elif section.startswith("nurse"):
                nurses.append({"id": section, **data})
            elif section.startswith("customer"):
                customers.append({"id": section, **data})
            elif section.startswith("job"):
                jobs.append({"id": section, **data})
        
        return normalization, nurses, customers, jobs

    def dispatch_caregivers(nurses,car_df,public_df):
        generated_caregiver = []

        for n in nurses:

            car_matrix_id = car_df.index.get_loc(n["id"])
            public_matrix_id = public_df.index.get_loc(n["id"])

            terminal_points.add(TerminalPoint(
                id=n["id"],
                location=n["home"],
                distance_matrix_index=car_matrix_id,
                public_distance_matrix_index=public_matrix_id
            ))   

            abilities = [n["qualification"]]
            for a in abilities:
                if a not in ["CSW","VN","HN","AHN","MN"]:
                    raise ValueError
                
                services.add(Service(
                            id=a,
                            type=a
                        ))
            
            aspects = n["aspects"].split(", ")
            if len(aspects) == 1:
                if aspects[0] == '':
                    aspects=None
            
            working_shift = (n["timewindow"][0] * tau,  n["timewindow"][1] * tau)

            nurse = Caregiver(
                id=n["id"],
                abilities=abilities,
                departing_point=n["id"],
                arrival_point=n["id"],
                aspects=aspects,
                transportation_mode=n["mode"],
                working_shift=working_shift
            )
            generated_caregiver.append(nurse)

        return generated_caregiver

    def dispatch_patients(customers, jobs, generated_caregivers, car_df_or, public_df_or, ):
        car_df = car_df_or.copy()
        public_df = public_df_or.copy()
        generated_patients = []

        for c in customers:
            required_services_tmp = c["jobs"].split(", ")
            aspects = c["aspects"].split(", ")
            if len(aspects) == 1:
                if aspects[0] == '':
                    aspects=None

            incompatible_caregivers = set()
            if aspects is not None:
                for caregiver in generated_caregivers:
                    caregiver_aspects = caregiver.aspects
                    if caregiver_aspects is None:
                        continue
                    if set(aspects).isdisjoint(set(caregiver_aspects)):
                        continue
                    incompatible_caregivers.add(caregiver.id)
            
            #syncronization = Synchronization(type="independent")
            
            for rs in required_services_tmp:
                # look for it in the list
                for j in jobs:
                    if j["id"] != rs:
                        continue
                    
                    preferred_start_time = j["preferredstarttime"] * tau
                    service_id = j["id"]
                    assert service_id == rs, "Mismatch between service customer and job (related to the service)"

                    service_type = j["qualification"]
                    patient_id = j["customer"]
                    time_window = (j["timewindow"][0] * tau, j["timewindow"][1] * tau)
                    duration = j["duration"] * tau

                    assert patient_id == c["id"], "Mismatch in the instances between customer and job (related to the customer)"

                    required_service = [RequiredService(
                        service=service_type,
                        duration=duration,
                    )]

                    services.add(Service(
                        id=service_type,
                        type=service_type
                    ))
                    car_matrix_id = car_df.index.get_loc(service_id)
                    public_matrix_id = public_df.index.get_loc(service_id)

                    preferred_caregivers = set()
                    for caregiver in generated_caregivers:
                        if caregiver.id in incompatible_caregivers:
                            continue
                        elif service_type in caregiver.abilities:
                            preferred_caregivers.add(caregiver.id)

                    patient = Patient(
                        id=f"{c['id']}-{service_id}",
                        required_services=required_service,
                        location=c["home"],
                        aspects=aspects,
                        time_windows=[time_window],
                        preferred_caregivers=preferred_caregivers,
                        preferred_start_time = preferred_start_time, 
                        # synchronization=syncronization,
                        incompatible_caregivers=incompatible_caregivers,
                        distance_matrix_index=car_matrix_id,
                        public_distance_matrix_index=public_matrix_id,
                        optional=True
                    )
                    generated_patients.append(patient)
                    break

        return generated_patients

    def dispatch_distance_matrix(distance_df):
        distances = []
        for i in range(distance_df.shape[0]):
            d = list(distance_df.iloc[i])
            distances.append(d)
        return distances

    normalization, nurses, customers, jobs = parse_entities_file(entities_file)
    car_matrix_df = pd.read_csv(car_matrix_file, sep=",",index_col=0)
    public_matrix_df = pd.read_csv(public_matrix_file, sep=",",index_col=0)

    car_matrix_df.columns = car_matrix_df.columns.str.strip()
    public_matrix_df.columns = public_matrix_df.columns.str.strip()

    def adjust_df_with_jobs(matrix_df, jobs, customers):
        for j in jobs:
            customer_identifier = j["customer"]
            job_identifier = j["id"]

            matrix_row = dict(matrix_df.loc[customer_identifier].copy())
            matrix_df.loc[job_identifier] = matrix_row
            
            matrix_col = list(matrix_df[customer_identifier].copy())

            matrix_df[job_identifier] = matrix_col
        matrix_df = matrix_df[~matrix_df.index.str.startswith('customer')]
        matrix_df = matrix_df.loc[:, ~matrix_df.columns.str.startswith('customer')]
        return matrix_df


    travel_time = 1.0
    # travel_time : normalization["travel_time"] = cost : normalization[cost]
    # cost = int(normalization[cost] * travel_time / normalization["travel_time"]) 
    qualification = int(normalization["qualification"] * travel_time / normalization["travel_time"]) 
    qualification = qualification if qualification > 0 else 1

    incompabilities = int(normalization["aspects"] * travel_time / normalization["travel_time"]) 
    incompabilities = incompabilities if incompabilities > 0 else 1

    total_extra_time = int(normalization["overtime"] * travel_time / normalization["travel_time"]) 
    total_extra_time = total_extra_time if total_extra_time > 0 else 1

    working_time = int(normalization["working_time"] * travel_time / normalization["travel_time"]) 
    working_time = working_time if working_time > 0 else 1

    tw_max_dev_in_desired_time = int(normalization["tw_max_dev_in_desired_time"] * travel_time / normalization["travel_time"]) 
    tw_max_dev_in_desired_time = tw_max_dev_in_desired_time if tw_max_dev_in_desired_time > 0 else 1

    tw_max_dev_in_time = int(normalization["tw_max_dev_in_time"] * travel_time / normalization["travel_time"]) 
    tw_max_dev_in_time = tw_max_dev_in_time if tw_max_dev_in_time > 0 else 1

    cost_components = CostComponents(
        tw_max_dev_in_time=tw_max_dev_in_time,
        tw_max_dev_in_desired_time=tw_max_dev_in_desired_time,
        working_time=working_time,
        total_extra_time=total_extra_time,
        travel_time=int(travel_time),
        incompabilities=incompabilities,
        qualification=qualification
    )
    
    car_matrix_df_modified = adjust_df_with_jobs(car_matrix_df,jobs,customers)
    public_matrix_df_modified = adjust_df_with_jobs(public_matrix_df,jobs,customers)

    metadata = MetaData(
        time_window_met="at_service_start", # FIXME: ask Luca
        origin="raidl",
        name=f"{output}",
        time_window_service_level=True,
        cost_components=cost_components
    )
    generated_caregivers = dispatch_caregivers(nurses=nurses, car_df=car_matrix_df, public_df=public_matrix_df)
    generated_patients = dispatch_patients(customers=customers, jobs=jobs, generated_caregivers=generated_caregivers, car_df_or=car_matrix_df_modified, public_df_or=public_matrix_df_modified,)
    car_matrix = [[(tau * element + 10) for element in row] for row in dispatch_distance_matrix(car_matrix_df_modified)]
    for i in range(len(car_matrix)):
        car_matrix[i][i] = 0
    public_matrix = [[(1.1 * tau * element + 5) for element in row] for row in dispatch_distance_matrix(public_matrix_df_modified)]
    for i in range(len(public_matrix)):
        public_matrix[i][i] = 0
    
    instance = Instance(
        metadata=metadata,
        patients=generated_patients,
        caregivers=generated_caregivers,
        distances=car_matrix,
        public_distances=public_matrix,
        terminal_points=terminal_points,
        services=list(services),
        services_order=["CSW","VN","HN","AHN","MN"])

    with open(f"{output}.json", "w") as f:
        f.write(instance.model_dump_json(indent=2, exclude_none=True))

@utils.command()
@click.argument('instance-filename', type=click.File())
@click.option('--origin', type=click.Choice(['kummer','extended', 'Italian', 'urli', 'mankowska', 'raidl', 'bazirha', 'bazirha-caie']), required=True, help="Original dataset of the instance")
@click.option('--output', '-o', type=click.File('w'), default='-', help="Output file")
def convert_instance(instance_filename, origin, output):
    try:
        i = Instance.model_validate_json(instance_filename.read())
        click.secho(f"Instance correctly converted OK", fg='green')
        
        if i.metadata.origin == 'oldies':
            warnings.warn(
                "'origin' not present in instance -- this is probably due to the fact you are converting an old instance so we are using the CLI field 'origin'",
                DeprecationWarning
            )  
            i.metadata.origin = origin
        
        if i.metadata.name == 'not-given' or i.metadata.origin == 'extended':
            warnings.warn(
                "'name' not present in instance -- this is probably due to the fact you are converting an old instansce so we are using its old name",
                DeprecationWarning
            ) 
            i.metadata.name = str(instance_filename.name).split(".json")[0].split("/")[-1]
        
        if not i.metadata.cost_components:
            warnings.warn(
                "'cost_components' not present in instance -- this is probably due to the fact you are converting an old instance so we are inferring the cost components using the CLU field 'origin'",
                DeprecationWarning
            ) 
            cost_components : CostComponents | None = None
            if origin in ["kummer", "mankowska","Italian"]: #FIXME: check what Italian is
                cost_components = CostComponents(
                    total_tardiness=1,
                    highest_tardiness=1,
                    travel_time=1
                )
            elif origin in ["extended"]:
                cost_components = CostComponents(
                    total_tardiness=1,
                    highest_tardiness=1,
                    travel_time=1,

                    total_waiting_time=1,
                    total_extra_time=1,
                    max_idle_time=1
                )
            else:
                # you should never end up here anyway
                raise ValueError("The 'origin' inferred from the CLI field does not have associated costs")
            assert cost_components is not None, "The field 'cost_component' should be not None"
            i.metadata.cost_components = cost_components
        
        output.write(i.model_dump_json(exclude_unset=True, indent=2))
    except ValidationError as e:
        click.secho(f"{e}", fg='red')

@utils.command()
@click.argument('instance-filename', type=click.File())
@click.argument('solution-filename', type=click.File())
@click.option('--output', '-o', type=click.File('w'), default='-', help="Output file")
def convert_solution(instance_filename, solution_filename, output):
    try:
        i = Instance.model_validate_json(instance_filename.read())
        s = Solution.model_validate_json(solution_filename.read())    
        click.secho(f"Validation OK, solution", fg='green')
        try:
            s.check_validity(i)
        except Exception as e:
            raise ValidationError(e)
        output.write(s.model_dump_json(exclude_unset=True, indent=2))
    except ValidationError as e:
        click.secho(f"{e}", fg='red')